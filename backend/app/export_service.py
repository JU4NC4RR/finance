"""
Genera copias de seguridad en Excel de los movimientos registrados.
Cada activo tiene su propio conjunto de columnas (las mismas que ve el
usuario en su formulario), mas los campos comunes id/fecha/nota, para que
el fichero sirva como copia de seguridad legible de lo que se ha ingresado.
"""

import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import models
from .backup_columns import columns_for_category, sanitize_sheet_name as _sanitize_sheet_name


def _write_sheet(ws: Worksheet, columns, transactions) -> None:
    ws.append([col.header for col in columns])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for tx in transactions:
        ws.append([col.to_cell(tx) for col in columns])
    for idx, col in enumerate(columns, start=1):
        width = max(len(col.header) + 2, 12)
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_asset_workbook(asset_type: models.AssetType, transactions: list[models.Transaction]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(asset_type.name)
    columns = columns_for_category(asset_type.category)
    _write_sheet(ws, columns, transactions)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_full_backup_workbook(asset_types: list[models.AssetType], transactions_by_asset: dict) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    for asset_type in asset_types:
        base_name = _sanitize_sheet_name(asset_type.name)
        sheet_name = base_name
        suffix = 2
        while sheet_name in used_names:
            sheet_name = _sanitize_sheet_name(f"{base_name} ({suffix})")
            suffix += 1
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        columns = columns_for_category(asset_type.category)
        _write_sheet(ws, columns, transactions_by_asset.get(asset_type.id, []))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
