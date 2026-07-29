"""
Restaura movimientos desde una copia de seguridad en Excel generada por
export_service. Es un "upsert": las filas con un ID que ya existe en la
base de datos se actualizan, las demas se crean. No se borra ningun
movimiento que no aparezca en el fichero, para que importar una copia
antigua nunca elimine datos mas recientes por accidente.
"""

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal

from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from . import models
from .backup_columns import columns_for_category, sanitize_sheet_name

REQUIRED_FIELDS = ("occurred_on", "quantity")
DEFAULTS = {"currency": "EUR", "purity": Decimal("1.0")}


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)
    unmatched_sheets: list[str] = field(default_factory=list)


def _read_rows(ws: Worksheet, columns):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        return
    header_index = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        values = {}
        for col in columns:
            idx = header_index.get(col.header)
            raw = row[idx] if idx is not None and idx < len(row) else None
            values[col.field] = col.from_cell(raw)
        yield values


def _derive_occurred_on(values: dict) -> None:
    # El efectivo solo exporta el mes del ingreso; se deriva el dia 1 de ese
    # mes para poder reconstruir occurred_on al reimportar.
    if values.get("occurred_on") is None and values.get("income_month"):
        try:
            values["occurred_on"] = date.fromisoformat(f"{values['income_month']}-01")
        except ValueError:
            pass


def _import_rows(db: Session, asset_type: models.AssetType, rows, result: ImportResult, sheet_label: str) -> None:
    columns = columns_for_category(asset_type.category)
    for row_num, values in enumerate(rows, start=2):
        _derive_occurred_on(values)
        missing = [f for f in REQUIRED_FIELDS if values.get(f) is None]
        if missing:
            result.skipped += 1
            result.errors.append(f"{sheet_label} fila {row_num}: faltan campos obligatorios ({', '.join(missing)})")
            continue

        for key, default in DEFAULTS.items():
            if key in {c.field for c in columns} and values.get(key) is None:
                values[key] = default

        tx_id = values.pop("id", None)
        created_at = values.pop("created_at", None)
        values.pop("asset_type_id", None)

        existing = db.get(models.Transaction, tx_id) if tx_id else None
        reuse_id = tx_id
        if existing is not None and existing.asset_type_id != asset_type.id:
            # el ID pertenece a un movimiento de otro activo: se crea uno nuevo con un ID distinto
            existing = None
            reuse_id = None

        if existing is not None:
            for key, value in values.items():
                setattr(existing, key, value)
            result.updated += 1
        else:
            obj = models.Transaction(asset_type_id=asset_type.id, **values)
            if reuse_id:
                obj.id = reuse_id
            if created_at:
                obj.created_at = created_at
            db.add(obj)
            result.created += 1


def import_asset_sheet(db: Session, asset_type: models.AssetType, ws: Worksheet) -> ImportResult:
    result = ImportResult()
    columns = columns_for_category(asset_type.category)
    rows = list(_read_rows(ws, columns))
    _import_rows(db, asset_type, rows, result, asset_type.name)
    db.commit()
    return result


def import_full_backup(db: Session, workbook, asset_types: list[models.AssetType]) -> ImportResult:
    result = ImportResult()
    by_sheet_name = {sanitize_sheet_name(a.name): a for a in asset_types}

    for sheet_name in workbook.sheetnames:
        asset_type = by_sheet_name.get(sheet_name)
        if asset_type is None:
            result.unmatched_sheets.append(sheet_name)
            continue
        ws = workbook[sheet_name]
        columns = columns_for_category(asset_type.category)
        rows = list(_read_rows(ws, columns))
        _import_rows(db, asset_type, rows, result, sheet_name)

    db.commit()
    return result
