import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from .. import crud, export_service, import_service, schemas
from ..database import get_db

router = APIRouter(prefix="/export", tags=["export"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _slug(name: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in name).strip("_") or "activo"


async def _load_workbook(file: UploadFile):
    content = await file.read()
    try:
        return load_workbook(io.BytesIO(content), data_only=True)
    except InvalidFileException:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel (.xlsx) valido")


@router.get("/full")
def export_full_backup(db: Session = Depends(get_db)):
    asset_types = crud.list_asset_types(db, only_active=False)
    transactions = crud.list_transactions(db, limit=100_000)

    transactions_by_asset: dict[uuid.UUID, list] = {}
    for tx in transactions:
        transactions_by_asset.setdefault(tx.asset_type_id, []).append(tx)

    buffer = export_service.build_full_backup_workbook(asset_types, transactions_by_asset)
    filename = f"boveda_backup_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/asset-types/{asset_type_id}")
def export_asset(asset_type_id: uuid.UUID, db: Session = Depends(get_db)):
    asset_type = crud.get_asset_type(db, asset_type_id)
    if asset_type is None:
        raise HTTPException(status_code=404, detail="asset_type_id no existe")

    transactions = crud.list_transactions(db, asset_type_id=asset_type_id, limit=100_000)
    buffer = export_service.build_asset_workbook(asset_type, transactions)
    filename = f"{_slug(asset_type.name)}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/full/import", response_model=schemas.ImportResultOut)
async def import_full_backup(file: UploadFile = File(...), db: Session = Depends(get_db)):
    workbook = await _load_workbook(file)
    asset_types = crud.list_asset_types(db, only_active=False)
    result = import_service.import_full_backup(db, workbook, asset_types)
    return schemas.ImportResultOut(**result.__dict__)


@router.post("/asset-types/{asset_type_id}/import", response_model=schemas.ImportResultOut)
async def import_asset(asset_type_id: uuid.UUID, file: UploadFile = File(...), db: Session = Depends(get_db)):
    asset_type = crud.get_asset_type(db, asset_type_id)
    if asset_type is None:
        raise HTTPException(status_code=404, detail="asset_type_id no existe")

    workbook = await _load_workbook(file)
    ws = workbook.active
    result = import_service.import_asset_sheet(db, asset_type, ws)
    return schemas.ImportResultOut(**result.__dict__)
